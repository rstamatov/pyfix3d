import vtk
import numpy as np
from matplotlib.path import Path
from chromosome_spline import *
import tkinter as tk
import os
from openpyxl import Workbook

class LineFitInteraction(vtk.vtkInteractorStyleTrackballCamera):

    def __init__(self, visualizer_3d, parent = None):

        self.visualizer_3d = visualizer_3d
        self.add_mouse_observers()
        
        self.ctrlPressed = False
        self.dragging = False
        self.clickedPos = None  

        self.cellLocator = vtk.vtkCellLocator()
        self.cellLocator.SetDataSet(self.visualizer_3d.imageData)
        self.cellLocator.BuildLocator()

        self.path = []  # To store the path of the drag
        self.polylineActor = vtk.vtkActor()
        self.polylineActor.GetProperty().SetColor(1.0, 1.0, 0.0)
        self.polylineActor.GetProperty().SetLineWidth(6)
        
        self.polylineMapper = vtk.vtkPolyDataMapper()
        self.polylineData = vtk.vtkPolyData()  # This will hold the points and lines
        self.polylinePoints = vtk.vtkPoints()
        self.polylineLines = vtk.vtkCellArray()

        self.polylineData.SetPoints(self.polylinePoints)
        self.polylineData.SetLines(self.polylineLines)

        self.polylineMapper.SetInputData(self.polylineData)
        self.polylineActor.SetMapper(self.polylineMapper)

        self.sphereActors = []

        self.visualizer_3d.renderer.AddActor(self.polylineActor)

        self.chromosome_models = {}

        for t in range(visualizer_3d.start_t, visualizer_3d.end_t + 1):
            self.chromosome_models[t] = {}

            for label in range(0, 256):
                self.chromosome_models[t][label] = None

    ##############################################################################################################################

    def add_mouse_observers(self):
        self.AddObserver("LeftButtonPressEvent", self.leftButtonPressEvent)
        self.AddObserver("RightButtonPressEvent", self.rightButtonPressEvent)
        self.AddObserver("LeftButtonReleaseEvent", self.leftButtonReleaseEvent)
        self.AddObserver("MouseMoveEvent", self.mouseMoveEvent)

    ##############################################################################################################################

    def remove_mouse_observers(self):
        self.RemoveAllObservers()

    ##############################################################################################################################

    def leftButtonPressEvent(self, obj, event):
                
        self.ctrlPressed = self.GetInteractor().GetControlKey()
        if self.ctrlPressed:
            self.clickedPos = self.GetInteractor().GetEventPosition()
            self.dragging = True

            self.choose_point(self.clickedPos)
            # Do not call OnLeftButtonDown to prevent camera rotation
        else:
            self.OnLeftButtonDown()

    ##############################################################################################################################

    def rightButtonPressEvent(self, obj, event):

        self.ctrlPressed = self.GetInteractor().GetControlKey()

        if self.ctrlPressed:
            self.fit_spline(smooth = 2, num_points = 100)
        else:
            self.clear_selection()

    ##############################################################################################################################


    def leftButtonReleaseEvent(self, obj, event):
                
        if self.ctrlPressed and self.dragging:
            
            self.endPos = self.GetInteractor().GetEventPosition()
            #self.markPointsWithinRectangle()
            self.dragging = False  # Reset dragging status
            #self.clearDrawnLine()
        else:
            self.OnLeftButtonUp()  # Handle normal release event

    ##############################################################################################################################
    
    def mouseMoveEvent(self, obj, event):
               
        self.OnMouseMove()

    ##############################################################################################################################

    def fit_spline(self, smooth, num_points):

        if len(self.path) == 0:
            return

        model = ChromosomeSpline(self.path,
                                 self.visualizer_3d.t,
                                 self.visualizer_3d.destination_color,
                                 smooth = 2,
                                 num_points = 100,
                                 spacing_z = self.visualizer_3d.spacing_z,
                                 spacing_y = self.visualizer_3d.spacing_y,
                                 spacing_x = self.visualizer_3d.spacing_x)
        
        self.chromosome_models[self.visualizer_3d.t][self.visualizer_3d.destination_color] = model

        actor = model.build_actor()
        
        # Add the actor to the existing renderer
        self.visualizer_3d.renderer.AddActor(actor)

        for actor in self.sphereActors:
            self.visualizer_3d.renderer.RemoveActor(actor)

        self.sphereActors = []
        self.path = []
        
        # Refresh the window to display the updates
        self.visualizer_3d.renderer.GetRenderWindow().Render()

    ##############################################################################################################################


    def choose_point(self, screen_coord):

        selected_object = self.visualizer_3d.destination_color

        if selected_object == 0:
            print("no object selected")
            return

        axisSpacing = self.visualizer_3d.imageData.GetSpacing()
        picker = self.visualizer_3d.picker

        xi, yi = screen_coord

        if picker.Pick(xi, yi, 0, self.visualizer_3d.renderer):
            pickedPosition = picker.GetPickPosition()

            rayStart = self.visualizer_3d.renderer.GetActiveCamera().GetPosition()
            rayEnd = pickedPosition

            rayDirection = np.subtract(rayEnd, rayStart)
            rayDirectionNormalized = rayDirection / np.linalg.norm(rayDirection)        

            rayEndExtended = pickedPosition + (rayDirectionNormalized * 100)
       

            # Surface object for the selected object or curve, if it's already built
            fitted_curve = self.chromosome_models[self.visualizer_3d.t][selected_object]
            if fitted_curve is not None:
                surfaceData = fitted_curve.actor.GetMapper().GetInput()
            else:
                surfaceData = self.visualizer_3d.marchingCubes[self.visualizer_3d.t][selected_object]

            # Setup vtkOBBTree for intersection test
            obbTree = vtk.vtkOBBTree()
            obbTree.SetDataSet(surfaceData)
            obbTree.BuildLocator()

            # Find intersection points
            points = vtk.vtkPoints()
            cellIds = vtk.vtkIdList()
            obbTree.IntersectWithLine(rayStart, rayEndExtended, points, cellIds)

            numPoints = points.GetNumberOfPoints()
            print  (numPoints)
            if numPoints > 1:
                # Assuming two intersection points (entry and exit), calculate midpoint
                p0 = np.array(points.GetPoint(0))
                p1 = np.array(points.GetPoint(numPoints-1)) # Last point in case of multiple intersections
                midpoint = (p0 + p1) / 2.0

                # Adjusting midpoint coordinates to world space if necessary
                midpoint_transformed = midpoint#[midpoint[d] * axisSpacing[d] for d in range(3)]

                self.path.append(midpoint_transformed)
                self.add_sphere_at_location(midpoint_transformed, 1, (1, 1, 0))

            elif numPoints == 1:
                # Handle the case with only one intersection point (e.g., ray touching surface)
                midpoint = np.array(points.GetPoint(0))
                midpoint_transformed = [midpoint[d] * axisSpacing[d] for d in range(3)]

                self.path.append(midpoint_transformed)
                self.add_sphere_at_location(midpoint_transformed, 1, (1, 1, 0))

            else:
                print("Ray does not intersect with the object.")
                return

        self.visualizer_3d.renderer.GetRenderWindow().Render()

    ###################################################################################################################################

    def add_sphere_at_location(self, location, radius, color):
        # Create a sphere source
        sphereSource = vtk.vtkSphereSource()
        sphereSource.SetCenter(location)
        sphereSource.SetRadius(radius)
        sphereSource.Update()

        # Create a mapper for the sphere source
        mapper = vtk.vtkPolyDataMapper()
        mapper.SetInputConnection(sphereSource.GetOutputPort())

        # Create an actor and set its mapper and color
        actor = vtk.vtkActor()
        actor.SetMapper(mapper)
        actor.GetProperty().SetColor(color)
        self.sphereActors.append(actor)
        self.visualizer_3d.renderer.AddActor(actor)

    ###################################################################################################################################

    def clear_selection(self):
        for actor in self.sphereActors:
            self.visualizer_3d.renderer.RemoveActor(actor)

        self.sphereActors = []
        self.path = []

        for model in self.chromosome_models[self.visualizer_3d.t].values():
            if model is not None:
                self.visualizer_3d.renderer.RemoveActor(model.actor)

        self.visualizer_3d.renderer.GetRenderWindow().Render()


    ###################################################################################################################################

    def hide_curves(self):
        for t in self.chromosome_models.keys():
            for model in self.chromosome_models[t].values():
                if model is not None:
                    self.visualizer_3d.renderer.RemoveActor(model.actor)

    ###################################################################################################################################

    def load_existing_models(self, clear = True):

        if clear:
            self.clear_selection()
        
        for label in self.chromosome_models[self.visualizer_3d.t].keys():

            if self.visualizer_3d.destination_color != 0 and label != self.visualizer_3d.destination_color:
                continue

            model = self.chromosome_models[self.visualizer_3d.t][label]
            if model is not None:
                self.visualizer_3d.renderer.AddActor(model.actor)

                if model.centromere_index > 0:
                    location = model.fitted_path[model.centromere_index]
                    self.add_sphere_at_location(location, 1, (1, 1, 0))
        
        self.visualizer_3d.renderer.GetRenderWindow().Render()

    ###################################################################################################################################

    def indicate_centromere(self):

        t = self.visualizer_3d.t
        label = self.visualizer_3d.destination_color
        curve = self.chromosome_models[t][label]
        
        if len(self.path) == 0 or curve is None:
            return
        
        point = self.path[0]

        # Find the closest point to the given, among all curves
        min_distance = np.Inf

            
        for i, p in enumerate(curve.fitted_path):
            current_distance = np.linalg.norm(np.array(point) - np.array(p))

            if current_distance < min_distance:
                min_distance = current_distance
                centromere_index = i

        curve.centromere_index = centromere_index
        arm_1, arm_2 = curve.measure_arms()

        ratio = arm_1 / (arm_1 + arm_2)
        if ratio > 0.5:
            ratio = 1 - ratio
        print ("p = ", arm_1)
        print ("q = ", arm_2)
        print ("ratio = ", ratio)

        self.load_existing_models()

    ###################################################################################################################################

    def save_current_splines(self):

        # Initiate a Tkinter root window but keep it hidden
        root = tk.Tk()
        root.withdraw()

        # Open a dialog to choose the save directory
        folder_selected = tk.filedialog.askdirectory()

        # Check if a folder was selected
        if not folder_selected:
            print("No folder selected.")
            return

        for t in self.chromosome_models.keys():
            for obj in self.chromosome_models[t].keys():                
        
                model = self.chromosome_models[t][obj]
                
                if model is not None:
                    model.save_to_file(folder_selected + "/model_" + str(t) + "_" + str(obj) + ".txt")
                    
        print ("Models saved.")

    ###################################################################################################################################

    def save_spline_measurements(self):

        # Get the desired file name and location for saving
        root = tk.Tk()
        root.withdraw()
        save_filename = tk.filedialog.asksaveasfilename(defaultextension = ".xlsx", filetypes = (("xlsx", "*.xlsx"), ("All Files", "*.*") ))
        
        # Create an empty excel sheet with tabs "Length" and "Ratio"
        workbook = Workbook()
        length_sheet = workbook.create_sheet("Length")
        ratio_sheet = workbook.create_sheet("Ratio")
        workbook.remove(workbook['Sheet'])  # Remove default sheet

        # Get the list of labels which have a at least one spline fit
        represented_labels = []
        for t in self.chromosome_models.keys():
            for obj in self.chromosome_models[t].keys():
                if self.chromosome_models[t][obj] is not None:
                    represented_labels.append(obj)
        represented_labels = list(np.unique(represented_labels))
    
        for t in self.chromosome_models.keys():

            # Row titles
            ratio_sheet.cell(row = 2 + t, column = 1, value = str(t))
            length_sheet.cell(row = 2 + t, column = 1, value = str(t))
            
            for i, obj in enumerate(represented_labels):                
                    
                model = self.chromosome_models[t][obj]

                if model is not None:


                    # Column title
                    length_sheet.cell(row = 1, column = 2 + i, value = "Chr " + str(obj))
                    ratio_sheet.cell(row = 1, column = 2 + i, value = "Chr " + str(obj))

                    arm_1, arm_2 = model.measure_arms()
                    length = arm_1 + arm_2

                    if arm_1 <= arm_2:
                       ratio = arm_1 / length
                    else:
                        ratio = arm_2 / length

                    # Save length in the "Length" tab at location (t, obj)
                    row_index = 2 + int(t)  # Edit according to your t values mapping to row
                    column_index = 2 + i
                    
                    # Save length in the "Length" tab at location (row_index, column_index)
                    length_sheet.cell(row = row_index, column = column_index, value = length)

                    # Save ratio in the "Ratio" tab at location (t, obj)
                    ratio_sheet.cell(row = row_index, column = column_index, value = ratio)

        workbook.save(filename = save_filename)
                

    ###################################################################################################################################

    def delete_active_spline(self):
        
        label = self.visualizer_3d.destination_color
        curve = self.chromosome_models[self.visualizer_3d.t][label]
        
        if curve is None:
            return

        # Remove this curve and update the view
        self.clear_selection()
        self.chromosome_models[self.visualizer_3d.t][label] = None
        self.load_existing_models(clear = False) # already cleared

    ###################################################################################################################################

    def load_from_file(self):

        # Initiate a Tkinter root window but keep it hidden
        root = tk.Tk()
        root.withdraw()

        # Open a dialog to choose the save directory
        folder_selected = tk.filedialog.askdirectory()

        # Check if a folder was selected
        if not folder_selected:
            print("No folder selected.")
            return

        for filename in os.listdir(folder_selected):

            if not filename.endswith(".txt"):
                continue

            points = []
            centromere_index = 0
            t = 0
            real_id = 0
            
            with open(folder_selected + "/" + filename, "r") as f:
                for line in f:
                    coords = line.split(" ")

                    if len(coords) == 3:
                        z, y, x = coords
                        points.append([float(z) * self.visualizer_3d.spacing_z,
                                       float(y) * self.visualizer_3d.spacing_y,
                                       float(x) * self.visualizer_3d.spacing_x])
                    else: # this is the centromere index
                        flag = coords[0]

                        if flag == "centromere":
                            centromere_index = int(coords[1])

                        elif flag == "time":
                            t = int(coords[1])

                        elif flag == "id":
                            real_id = float(coords[1])
                            

            model = ChromosomeSpline(points,
                                     t,
                                     real_id,
                                     smooth = 0,
                                     num_points = 100,
                                     spacing_z = self.visualizer_3d.spacing_z,
                                     spacing_y = self.visualizer_3d.spacing_y,
                                     spacing_x = self.visualizer_3d.spacing_x)

            model.build_actor()
            model.centromere_index = centromere_index
            
            self.chromosome_models[t][real_id] = model

        self.load_existing_models()
                    
    

    

    
